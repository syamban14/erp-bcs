try:
    import pandas as pd
    df = pd.read_excel('jadwalKerjaRoster.xlsx')
    print("Columns:", df.columns.tolist())
    print("First 5 rows:")
    print(df.head(5))
except ImportError:
    print("pandas not found, trying openpyxl")
    try:
        from openpyxl import load_workbook
        wb = load_workbook('jadwalKerjaRoster.xlsx')
        ws = wb.active
        print("Rows:")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 5: break
            print(row)
    except ImportError:
        print("openpyxl not found")
except Exception as e:
    print("Error:", e)
